"""HTTP-обробники (views) для модуля управління ремонтами.

Модуль містить усі view-функції для роботи з:
- заявками на ремонт (Repair): список, деталі, створення, редагування, видалення
- клієнтами (Client): список, деталі, створення, редагування
- пристроями (Device): створення
- запчастинами (Part): видалення
- службовими ендпоінтами: AJAX для пристроїв клієнта, експорт Excel

Взаємодія компонентів::

    URL Router → View → Service (бізнес-логіка) → Model (БД)
                  ↓
               Template (HTML)

Усі views захищені ``@login_required``.
Бізнес-логіка (переходи статусів, обмеження полів) делегована в ``services.py``.
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone

from .models import Repair, Client, Device, Part
from .forms import (
    RepairForm,
    RepairUpdateForm,
    ClientForm,
    DeviceForm,
    RepairCommentForm,
    PartForm,
    RepairFilterForm,
)
from .services import update_repair, RepairTransitionError

# Список міст України для автодоповнення у формі клієнта
UA_CITIES = [
    "Київ",
    "Харків",
    "Одеса",
    "Дніпро",
    "Запоріжжя",
    "Львів",
    "Вінниця",
    "Кривий Ріг",
    "Миколаїв",
    "Полтава",
    "Чернігів",
    "Черкаси",
    "Суми",
    "Житомир",
    "Рівне",
    "Івано-Франківськ",
    "Тернопіль",
    "Ужгород",
    "Хмельницький",
    "Чернівці",
    "Кропивницький",
    "Луцьк",
    "Біла Церква",
    "Бровари",
    "Ірпінь",
    "Кременчук",
]


@login_required
def dashboard(request):
    """Відображає головний дашборд сервісного центру.

    Збирає зведену статистику у вигляді словника ``stats``:
    - загальна кількість заявок
    - нові заявки (очікують призначення)
    - заявки в роботі
    - виконані
    - прострочені (дедлайн минув, заявка відкрита)
    - мої заявки (активні для поточного користувача)

    Також формує два списки:
    - ``recent_repairs`` — 10 найновіших заявок
    - ``urgent_repairs`` — 5 активних заявок з пріоритетом high/urgent

    Args:
        request: Django HttpRequest об'єкт.

    Returns:
        HttpResponse зі шаблоном ``repairs/dashboard.html``.
    """
    repairs = Repair.objects.select_related("client", "master", "device")
    today = timezone.now().date()
    stats = {
        "total": repairs.count(),
        "new": repairs.filter(status="new").count(),
        "in_progress": repairs.filter(status="in_progress").count(),
        "done": repairs.filter(status="done").count(),
        "overdue": repairs.filter(deadline__lt=today)
        .exclude(status__in=Repair.CLOSED_STATUSES)
        .count(),
        "my_repairs": repairs.filter(master=request.user)
        .exclude(status__in=Repair.CLOSED_STATUSES)
        .count(),
    }
    recent_repairs = repairs.order_by("-created_at")[:10]
    urgent_repairs = (
        repairs.filter(priority__in=["high", "urgent"])
        .exclude(status__in=Repair.CLOSED_STATUSES)
        .order_by("-created_at")[:5]
    )
    return render(
        request,
        "repairs/dashboard.html",
        {
            "stats": stats,
            "recent_repairs": recent_repairs,
            "urgent_repairs": urgent_repairs,
        },
    )


@login_required
def repair_list(request):
    """Відображає список заявок на ремонт з фільтрацією.

    Підтримує фільтрацію через GET-параметри форми ``RepairFilterForm``:
    - ``search`` — пошук по номеру, імені/прізвищу клієнта, телефону, опису
    - ``status`` — фільтр за статусом
    - ``priority`` — фільтр за пріоритетом
    - ``master`` — фільтр за майстром
    - ``date_from`` / ``date_to`` — діапазон дат створення

    Args:
        request: Django HttpRequest з GET-параметрами фільтрів.

    Returns:
        HttpResponse зі шаблоном ``repairs/repair_list.html``.
    """
    form = RepairFilterForm(request.GET)
    repairs = Repair.objects.select_related("client", "master", "device").order_by("-created_at")
    if form.is_valid():
        if form.cleaned_data.get("search"):
            q = form.cleaned_data["search"]
            repairs = repairs.filter(
                Q(number__icontains=q)
                | Q(client__first_name__icontains=q)
                | Q(client__last_name__icontains=q)
                | Q(client__phone__icontains=q)
                | Q(problem_description__icontains=q)
            )
        if form.cleaned_data.get("status"):
            repairs = repairs.filter(status=form.cleaned_data["status"])
        if form.cleaned_data.get("priority"):
            repairs = repairs.filter(priority=form.cleaned_data["priority"])
        if form.cleaned_data.get("master"):
            repairs = repairs.filter(master=form.cleaned_data["master"])
        if form.cleaned_data.get("date_from"):
            repairs = repairs.filter(created_at__date__gte=form.cleaned_data["date_from"])
        if form.cleaned_data.get("date_to"):
            repairs = repairs.filter(created_at__date__lte=form.cleaned_data["date_to"])
    return render(
        request,
        "repairs/repair_list.html",
        {"repairs": repairs, "form": form, "total": repairs.count()},
    )


@login_required
def repair_detail(request, pk):
    """Відображає деталі заявки та обробляє додавання коментарів/запчастин.

    GET: повертає сторінку деталей з формами коментаря та запчастини.

    POST обробляє два сценарії за кнопкою submit:
    - ``add_comment`` — додає внутрішній коментар до заявки
    - ``add_part`` — додає запчастину (тільки якщо статус дозволяє)

    Закриті заявки (done/issued/cancelled) не приймають POST-запити —
    повертається попередження і redirect.

    Args:
        request: Django HttpRequest об'єкт.
        pk: Первинний ключ заявки Repair.

    Returns:
        HttpResponse зі шаблоном ``repairs/repair_detail.html`` або redirect.
    """
    repair = get_object_or_404(
        Repair.objects.select_related("client", "master", "device", "created_by"), pk=pk
    )
    comment_form = RepairCommentForm()
    part_form = PartForm()
    parts_allowed = repair.can_add_parts()

    if request.method == "POST":
        if repair.is_closed():
            messages.warning(request, "Заявка закрита — зміни неможливі.")
            return redirect("repair_detail", pk=pk)

        if "add_comment" in request.POST:
            comment_form = RepairCommentForm(request.POST)
            if comment_form.is_valid():
                c = comment_form.save(commit=False)
                c.repair = repair
                c.author = request.user
                c.save()
                messages.success(request, "Коментар додано.")
                return redirect("repair_detail", pk=pk)

        elif "add_part" in request.POST:
            if not parts_allowed:
                messages.error(
                    request,
                    "Запчастини можна додавати лише при статусах: "
                    "Діагностовано, В роботі, Очікування запчастин.",
                )
                return redirect("repair_detail", pk=pk)
            part_form = PartForm(request.POST)
            if part_form.is_valid():
                p = part_form.save(commit=False)
                p.repair = repair
                p.save()
                messages.success(request, "Запчастину додано.")
                return redirect("repair_detail", pk=pk)

    parts = repair.parts.all()
    parts_total = sum(p.total() for p in parts)
    labor_cost = repair.labor_cost or 0
    return render(
        request,
        "repairs/repair_detail.html",
        {
            "repair": repair,
            "comment_form": comment_form,
            "part_form": part_form,
            "parts": parts,
            "comments": repair.comments.select_related("author").all(),
            "parts_total": parts_total,
            "labor_cost": labor_cost,
            "total_cost": parts_total + labor_cost,
            "parts_allowed": parts_allowed,
        },
    )


@login_required
def repair_create(request):
    """Обробляє створення нової заявки на ремонт.

    GET: повертає порожню форму ``RepairForm`` зі списком клієнтів
    для вибору (передається окремо для JS-автодоповнення пристроїв).

    POST: валідує форму, встановлює ``created_by = request.user``,
    зберігає заявку та перенаправляє на її сторінку деталей.

    Args:
        request: Django HttpRequest об'єкт.

    Returns:
        HttpResponse зі шаблоном ``repairs/repair_form.html`` або
        redirect на ``repair_detail`` після успішного створення.
    """
    if request.method == "POST":
        form = RepairForm(request.POST)
        if form.is_valid():
            repair = form.save(commit=False)
            repair.created_by = request.user
            repair.save()
            messages.success(request, f"Заявку {repair.number} створено!")
            return redirect("repair_detail", pk=repair.pk)
    else:
        form = RepairForm()
    return render(
        request,
        "repairs/repair_form.html",
        {
            "form": form,
            "title": "Нова заявка",
            "action": "Створити заявку",
            "clients": Client.objects.order_by("last_name", "first_name"),
        },
    )


@login_required
def repair_edit(request, pk):
    """Обробляє редагування існуючої заявки на ремонт.

    Видані заявки (статус ``issued``) не підлягають редагуванню —
    повертається помилка і redirect на деталі.

    POST: делегує оновлення в ``services.update_repair()``, який
    виконує валідацію переходів та обмеження полів. При
    ``RepairTransitionError`` повертає форму з повідомленням про помилку.

    Args:
        request: Django HttpRequest об'єкт.
        pk: Первинний ключ заявки Repair.

    Returns:
        HttpResponse зі шаблоном ``repairs/repair_form.html`` або
        redirect на ``repair_detail`` після успішного збереження.
    """
    repair = get_object_or_404(Repair, pk=pk)

    if repair.status == Repair.STATUS_ISSUED:
        messages.error(request, "Видана заявка не підлягає редагуванню.")
        return redirect("repair_detail", pk=pk)

    if request.method == "POST":
        form = RepairUpdateForm(request.POST, instance=repair, current_user=request.user)
        if form.is_valid():
            try:
                update_repair(
                    repair=repair,
                    data=form.cleaned_data,
                    user=request.user,
                )
                messages.success(request, "Заявку оновлено.")
                return redirect("repair_detail", pk=pk)
            except RepairTransitionError as exc:
                messages.error(request, str(exc))
    else:
        form = RepairUpdateForm(instance=repair, current_user=request.user)

    return render(
        request,
        "repairs/repair_form.html",
        {
            "form": form,
            "repair": repair,
            "title": f"Редагування {repair.number}",
            "action": "Зберегти зміни",
            "is_create": False,
        },
    )


@login_required
def repair_delete(request, pk):
    """Обробляє видалення заявки на ремонт з підтвердженням.

    Видалення заборонене для заявок зі статусами ``done`` та ``issued``,
    щоб зберегти історію завершених ремонтів.

    GET: повертає сторінку підтвердження видалення.
    POST: виконує видалення та перенаправляє на список заявок.

    Args:
        request: Django HttpRequest об'єкт.
        pk: Первинний ключ заявки Repair.

    Returns:
        HttpResponse зі шаблоном ``repairs/repair_confirm_delete.html``
        або redirect на ``repair_list`` після видалення.
    """
    repair = get_object_or_404(Repair, pk=pk)
    if repair.status in [Repair.STATUS_ISSUED, Repair.STATUS_DONE]:
        messages.error(
            request, f"Заявку зі статусом «{repair.get_status_display()}» видалити неможливо."
        )
        return redirect("repair_detail", pk=pk)
    if request.method == "POST":
        number = repair.number
        repair.delete()
        messages.success(request, f"Заявку {number} видалено.")
        return redirect("repair_list")
    return render(request, "repairs/repair_confirm_delete.html", {"repair": repair})


@login_required
def client_list(request):
    """Відображає список клієнтів з пошуком.

    Підтримує текстовий пошук по імені, прізвищу, телефону та email
    через GET-параметр ``search``. Додає анотацію ``repair_count`` —
    загальну кількість заявок клієнта.

    Args:
        request: Django HttpRequest з опційним GET параметром ``search``.

    Returns:
        HttpResponse зі шаблоном ``repairs/client_list.html``.
    """
    search = request.GET.get("search", "")
    clients = Client.objects.annotate(repair_count=Count("repairs")).order_by("-created_at")
    if search:
        clients = clients.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(phone__icontains=search)
            | Q(email__icontains=search)
        )
    return render(request, "repairs/client_list.html", {"clients": clients, "search": search})


@login_required
def client_create(request):
    """Обробляє створення нового клієнта.

    Args:
        request: Django HttpRequest об'єкт.

    Returns:
        HttpResponse зі шаблоном ``repairs/client_form.html`` або
        redirect на ``client_list`` після успішного збереження.
    """
    if request.method == "POST":
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save()
            messages.success(request, f"Клієнта {client.get_full_name()} додано.")
            return redirect("client_list")
    else:
        form = ClientForm()
    return render(
        request,
        "repairs/client_form.html",
        {"form": form, "title": "Новий клієнт", "cities": UA_CITIES},
    )


@login_required
def client_edit(request, pk):
    """Обробляє редагування даних існуючого клієнта.

    Args:
        request: Django HttpRequest об'єкт.
        pk: Первинний ключ клієнта Client.

    Returns:
        HttpResponse зі шаблоном ``repairs/client_form.html`` або
        redirect на ``client_detail`` після успішного збереження.
    """
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, "Дані клієнта оновлено.")
            return redirect("client_detail", pk=pk)
    else:
        form = ClientForm(instance=client)
    return render(
        request,
        "repairs/client_form.html",
        {
            "form": form,
            "client": client,
            "title": f"Редагування: {client.get_full_name()}",
            "cities": UA_CITIES,
        },
    )


@login_required
def client_detail(request, pk):
    """Відображає картку клієнта з його пристроями та заявками.

    Args:
        request: Django HttpRequest об'єкт.
        pk: Первинний ключ клієнта Client.

    Returns:
        HttpResponse зі шаблоном ``repairs/client_detail.html``.
    """
    client = get_object_or_404(Client, pk=pk)
    repairs = client.repairs.select_related("master", "device").order_by("-created_at")
    return render(
        request,
        "repairs/client_detail.html",
        {"client": client, "repairs": repairs, "devices": client.devices.all()},
    )


@login_required
def device_create(request, client_pk):
    """Обробляє додавання нового пристрою до облікового запису клієнта.

    Args:
        request: Django HttpRequest об'єкт.
        client_pk: Первинний ключ клієнта якому додається пристрій.

    Returns:
        HttpResponse зі шаблоном ``repairs/device_form.html`` або
        redirect на ``client_detail`` після успішного збереження.
    """
    client = get_object_or_404(Client, pk=client_pk)
    if request.method == "POST":
        form = DeviceForm(request.POST)
        if form.is_valid():
            device = form.save(commit=False)
            device.client = client
            device.save()
            messages.success(request, "Пристрій додано.")
            return redirect("client_detail", pk=client_pk)
    else:
        form = DeviceForm()
    return render(request, "repairs/device_form.html", {"form": form, "client": client})


@login_required
def delete_part(request, pk):
    """Видаляє запчастину із заявки.

    Видалення можливе лише якщо поточний статус заявки дозволяє
    редагування запчастин (``repair.can_add_parts()``).

    GET та некоректні POST: перенаправляє на деталі заявки без змін.
    POST: видаляє запчастину та перенаправляє на деталі заявки.

    Args:
        request: Django HttpRequest об'єкт.
        pk: Первинний ключ запчастини Part.

    Returns:
        Redirect на ``repair_detail`` відповідної заявки.
    """
    part = get_object_or_404(Part, pk=pk)
    repair = part.repair
    if not repair.can_add_parts():
        messages.error(request, "Не можна видаляти запчастини при поточному статусі заявки.")
        return redirect("repair_detail", pk=repair.pk)
    if request.method == "POST":
        repair_pk = repair.pk
        part.delete()
        messages.success(request, "Запчастину видалено.")
        return redirect("repair_detail", pk=repair_pk)
    return redirect("repair_detail", pk=repair.pk)


@login_required
def get_client_devices(request):
    """AJAX-ендпоінт: повертає список пристроїв клієнта у форматі JSON.

    Використовується на сторінці створення заявки для динамічного
    оновлення списку пристроїв при зміні клієнта (без перезавантаження).

    GET параметри:
        client_id (int): ID клієнта чиї пристрої треба отримати.

    Returns:
        JsonResponse зі списком об'єктів ``[{id, name}, ...]``,
        де ``name`` = ``"{brand} {model}"``.
        Повертає порожній список якщо client_id невалідний.

    Example:
        GET /repairs/client-devices/?client_id=5
        → [{"id": 3, "name": "Apple iPhone 13"}, ...]
    """
    try:
        client_id = int(request.GET.get("client_id", 0))
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    devices = Device.objects.filter(client_id=client_id).values("id", "brand", "model")
    return JsonResponse(
        [{"id": d["id"], "name": f"{d['brand']} {d['model']}"} for d in devices], safe=False
    )


@login_required
def export_repairs_excel(request):
    """Експортує всі заявки на ремонт у форматі Excel (.xlsx).

    Генерує файл за допомогою openpyxl з:
    - заголовком з синім фоном та білим текстом
    - тонкими рамками для кожної клітинки
    - фіксованою шириною колонок

    Колонки: №, Номер, Клієнт, Телефон, Місто, Пристрій, Проблема,
    Статус, Пріоритет, Майстер, Запчастини (грн), Робота (грн),
    Разом (грн), Дата прийому, Дедлайн, Виконано.

    Returns:
        HttpResponse з Content-Type ``application/vnd.openxmlformats...``
        та заголовком Content-Disposition для завантаження файлу.
        При відсутності openpyxl — redirect на ``repair_list`` з помилкою.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "Бібліотека openpyxl не встановлена.")
        return redirect("repair_list")

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfil = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "№",
        "Номер",
        "Клієнт",
        "Телефон",
        "Місто",
        "Пристрій",
        "Проблема",
        "Статус",
        "Пріоритет",
        "Майстер",
        "Запчастини (грн)",
        "Робота (грн)",
        "Разом (грн)",
        "Дата прийому",
        "Дедлайн",
        "Виконано",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hf
        cell.fill = hfil
        cell.alignment = ca
        cell.border = tb

    sm = dict(Repair.STATUS_CHOICES)
    pm = dict(Repair.PRIORITY_CHOICES)
    repairs = Repair.objects.select_related("client", "master", "device").order_by("-created_at")

    for i, r in enumerate(repairs, 1):
        pt = float(r.parts_total())
        lc = float(r.labor_cost or 0)
        ws.append(
            [
                i,
                r.number,
                r.client.get_full_name(),
                r.client.phone,
                r.client.city or "—",
                str(r.device) if r.device else "—",
                r.problem_description[:80],
                sm.get(r.status, r.status),
                pm.get(r.priority, r.priority),
                r.master.get_full_name() if r.master else "—",
                pt,
                lc,
                pt + lc,
                r.created_at.strftime("%d.%m.%Y %H:%M"),
                r.deadline.strftime("%d.%m.%Y") if r.deadline else "—",
                r.completed_at.strftime("%d.%m.%Y") if r.completed_at else "—",
            ]
        )
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i + 1, column=col)
            cell.border = tb
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for i, w in enumerate([5, 13, 22, 14, 14, 22, 30, 16, 13, 20, 15, 13, 13, 18, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="repairs_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response
